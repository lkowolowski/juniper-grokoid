#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl>=3.0.0",
# ]
# ///
"""juniper_grokoid — Junos Syslog → Graylog Grok Extractor Generator.

Reads Juniper Junos OS syslog message definitions from an Excel spreadsheet
and generates Graylog-compatible Grok extractor JSON.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import warnings

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CATEGORY_MAP = {
    "security": [
        "RT_SCREEN_",
        "RT_IPSEC_",
        "IDP_",
        "KMD_",
        "IKE_",
        "IPSEC_",
        "FWAUTH_",
        "SSL_PROXY_",
        "SECINTEL_",
        "PFE_FW_",
        "PFE_SCREEN_",
        "DDOS_",
    ],
    "utm": [
        "WEBFILTER_",
        "AAMW_",
        "AV_",
        "ANTI_VIRUS_",
        "ANTISPAM_",
        "URLFD_",
        "UTMD_",
        "ICAP_",
        "APPTRACK_",
    ],
    "flow": [
        "RT_FLOW_",
        "RT_ALG_",
        "RT_GTP_",
        "RT_SCTP_",
        "RT_NAT_",
        "RT_SRC_",
        "RT_DST_",
        "RT_STATIC_",
        "FLOW_",
        "PFE_FLOWD_",
    ],
    "chassis": ["CHASSISD_"],
    "routing": [
        "BGP_",
        "RPD_OSPF_",
        "RPD_ISIS_",
        "RPD_BGP_",
        "RPD_LDP_",
        "RPD_MPLS_",
        "RPD_RSVP_",
        "RPD_PIM_",
        "RPD_IGMP_",
        "RPD_MLD_",
        "RPD_KRT_",
        "RPD_RT_",
        "RPD_SPRING_",
        "EVPN_",
        "MPLS_",
    ],
}

VARIABLE_RE = re.compile(r"<variable>(.*?)</variable>")

# ---------------------------------------------------------------------------
# Pure functions
# ---------------------------------------------------------------------------


def classify_variable(name: str) -> str:
    """Classify a variable name as IP, NUMBER, or DATA."""
    lower = name.lower()
    if ("address" in lower or "ip" in lower) and "port" not in lower:
        return "IP"
    for kw in ("port", "count", "packets", "bytes", "elapsed", "number"):
        if kw in lower:
            return "NUMBER"
    if name == "id" or name.endswith("-id"):
        return "NUMBER"
    return "DATA"


def variable_to_grok(name: str) -> str:
    """Convert a variable name to a Grok capture, e.g. ``%{IP:source_address}``."""
    vtype = classify_variable(name)
    field = name.replace("-", "_")
    return f"%{{{vtype}:{field}}}"


def template_to_grok(template: str) -> str:
    """Convert a Junos syslog message template to a Grok pattern.

    ``<variable>…</variable>`` tags become Grok captures; literal text is
    regex-escaped so that characters like ``(``, ``)``, etc. are matched
    literally.
    """
    template = template.rstrip()
    parts = re.split(VARIABLE_RE, template)
    result = []
    for i, part in enumerate(parts):
        if i % 2 == 0:
            result.append(re.escape(part))
        else:
            result.append(variable_to_grok(part))
    output = "".join(result)
    assert "<variable>" not in output, "Unconverted <variable> tag in output"
    return output


def extract_category(name: str) -> str | None:
    """Return the category for a syslog message *name*, or ``None``."""
    for category, prefixes in CATEGORY_MAP.items():
        for prefix in prefixes:
            if name.startswith(prefix):
                return category
    return None


def has_variables(message: str) -> bool:
    """Return whether *message* contains ``<variable>`` tags."""
    return "<variable>" in message


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------


def parse_excel(filepath: str) -> list[dict]:
    """Parse the Junos syslog Excel file and return a list of dicts.

    Each dict has keys ``"name"`` and ``"message"``.
    """
    warnings.filterwarnings("ignore", message=".*no default style.*")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find the right sheet
    target_sheet = "System Log Messages"
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
    else:
        ws = wb.worksheets[0]
        print(
            f"Warning: sheet '{target_sheet}' not found, using '{ws.title}'",
            file=sys.stderr,
        )

    # Auto-detect header row: scan first 50 rows for NAME + MESSAGE
    header_row = None
    name_col = None
    msg_col = None
    for row_idx in range(1, min(51, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().upper() == "NAME":
                name_col = col_idx
            if val and str(val).strip().upper() == "MESSAGE":
                msg_col = col_idx
        if name_col and msg_col:
            header_row = row_idx
            break

    if header_row is None:
        wb.close()
        raise ValueError("Could not find header row with NAME and MESSAGE columns")

    # Read data rows after header
    seen = set()
    records = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        message = ws.cell(row=row_idx, column=msg_col).value
        if not name:
            continue
        name = str(name).strip()
        message = str(message).strip() if message else ""
        if name in seen:
            continue
        seen.add(name)
        records.append({"name": name, "message": message})

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Output building
# ---------------------------------------------------------------------------


def build_extractor(name: str, grok_pattern: str, order: int) -> dict:
    """Build a single Graylog extractor dict."""
    return {
        "title": name,
        "extractor_type": "grok",
        "converters": [],
        "order": order,
        "cursor_strategy": "copy",
        "source_field": "message",
        "target_field": "",
        "extractor_config": {"grok_pattern": grok_pattern},
        "condition_type": "string",
        "condition_value": name,
    }


def build_extractors_json(extractors: list[dict], version: str) -> dict:
    """Wrap extractor dicts in the top-level Graylog JSON envelope."""
    return {"extractors": extractors, "version": version}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv=None):
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        prog="juniper_grokoid",
        description="Generate Graylog Grok extractors from Junos syslog definitions.",
    )
    parser.add_argument(
        "--input",
        "-i",
        help="Path to the Junos syslog Excel file.",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Output JSON file (default: stdout).",
    )
    parser.add_argument(
        "--categories",
        "-c",
        help="Comma-separated list of categories to include (default: all).",
    )
    parser.add_argument(
        "--dry-run",
        "-n",
        action="store_true",
        help="Print name + grok pattern pairs instead of JSON.",
    )
    parser.add_argument(
        "--list-categories",
        action="store_true",
        help="Print available categories and exit.",
    )
    parser.add_argument(
        "--stats",
        action="store_true",
        help="Show counts by category.",
    )
    parser.add_argument(
        "--graylog-version",
        default="5.0.13",
        help="Graylog version string (default: 5.0.13).",
    )

    args = parser.parse_args(argv)

    if not args.list_categories and not args.input:
        parser.error("--input / -i is required (unless using --list-categories)")

    # Parse comma-separated categories
    if args.categories:
        args.categories = [c.strip() for c in args.categories.split(",")]

    return args


def main():
    args = parse_args()

    # --list-categories -------------------------------------------------------
    if args.list_categories:
        for cat, prefixes in CATEGORY_MAP.items():
            print(f"{cat}:")
            for p in prefixes:
                print(f"  {p}")
        return

    # Parse Excel -------------------------------------------------------------
    records = parse_excel(args.input)

    # Filter & generate -------------------------------------------------------
    selected_categories = args.categories
    extractors = []
    category_counts: dict[str, int] = {}
    order = 0

    for rec in records:
        name = rec["name"]
        message = rec["message"]

        if not message or not has_variables(message):
            continue

        category = extract_category(name)
        if category is None:
            continue

        if selected_categories and category not in selected_categories:
            continue

        grok_pattern = template_to_grok(message)
        ext = build_extractor(name, grok_pattern, order)
        extractors.append(ext)
        category_counts[category] = category_counts.get(category, 0) + 1
        order += 1

    # --dry-run ---------------------------------------------------------------
    if args.dry_run:
        for ext in extractors:
            print(f"{ext['title']}  {ext['extractor_config']['grok_pattern']}")
        return

    # --stats -----------------------------------------------------------------
    if args.stats:
        for cat, count in sorted(category_counts.items()):
            print(f"{cat}: {count}")
        print(f"TOTAL: {sum(category_counts.values())}")
        return

    # Write JSON output -------------------------------------------------------
    output = build_extractors_json(extractors, args.graylog_version)
    json_str = json.dumps(output, indent=2)

    if args.output:
        with open(args.output, "w") as f:
            f.write(json_str + "\n")
    else:
        print(json_str)

    # Summary to stderr -------------------------------------------------------
    parts = [f"{count} {cat}" for cat, count in sorted(category_counts.items())]
    dest = args.output if args.output else "stdout"
    print(
        f"Generated {len(extractors)} extractors ({', '.join(parts)}) → {dest}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
